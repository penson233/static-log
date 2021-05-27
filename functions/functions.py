# -*- coding: utf-8 -*-
# @File   : hello
# @Author : penson <penson@penson.top>
# @Email: decentpenson@gmail.com
# @Date   : 2021/2/19 11:40

import os
import re
import time
from openpyxl import Workbook
from collections import Counter
import requests
from prettytable import PrettyTable
from functions.defend import blacklist_find

# 得到文件创建时间
def get_filetime(file):
    times = os.path.getmtime(file)
    file_time = time.strftime("%Y-%h-%d", time.localtime(times))
    return file_time

#创建目录
def is_path(path):
    isExists = os.path.exists(path)

    if not isExists:
        os.makedirs(path)

    else:
        return None

#保存为表格
def savetable(table,ip,country,province,city,type,count):
    table.add_row([ip,country,province,city,type,count])
    print(table)

#查找日志文件
def find_log(path):
    files = os.listdir(path)
    dir_list = sorted(files,key=lambda x: os.path.getmtime(os.path.join(path,x)))
    return dir_list

#找到安全攻击
def find_attack(req,attack_json,ip):
    if blacklist_find(req[0]) == "sql":
        attack_json.append([ip[0], "sql注入", "中危", req[0]])
    elif blacklist_find(req[0]) == "xss":
        attack_json.append([ip[0], "xss攻击", "中危", req[0]])
    elif blacklist_find(req[0]) == "ssrf":
        attack_json.append([ip[0], "ssrf攻击", "高危", req[0]])
    elif blacklist_find(req[0]) == "php":
        attack_json.append([ip[0], "php高危函数", "中危", req[0]])
    elif blacklist_find(req[0]) == "brute" and req[:-1] == "200":
        attack_json.append([ip[0], "敏感文件泄露", "低危", req[0]])

def save_excel(save_path, file_time, outjson,outjson_ua,ip_list):


    # 保存请求
    wb = Workbook()
    ws = wb.active

    # 保存ua
    wb_ua = Workbook()
    ws_ua = wb_ua.active

    #保存攻击
    attack_wb= Workbook()
    attack_ws = attack_wb.active


    attack_json=[]



    ws.append(["ip", "请求", "响应状态码"])
    ws_ua.append(["ip", "请求","请求头"])
    attack_ws.append(["ip","攻击类型","风险","请求"])


    for result in outjson["result"]:
        for ip in ip_list:
            if ip[0] in result:
                for i in result[ip[0]]:
                    req = str(i).replace('{', '').replace('}', '').replace('\'', '').split(':')
                    find_attack(req,attack_json,ip)
                    if len(req) > 2:
                        ws.append([ip[0], req[0], req[2]])
                    else:
                        ws.append([ip[0], req[0], req[1]])

    for ua in outjson_ua["result"]:
        ws_ua.append([ua[0],ua[1],ua[2]])

    for i in attack_json:
        attack_ws.append([i[0],i[1],i[2],i[3]])

    is_path(f"{save_path}\{file_time}")
    wb.save(f"{save_path}\{file_time}\{file_time}.xls")
    wb_ua.save(f"{save_path}\{file_time}\\ua.xls")
    attack_wb.save(f"{save_path}\{file_time}\\攻击统计.xls")



#找到日志文件中的ip，请求方式，请求方式的次数，访问的次数
def finding(text,file_time,save_path):

    is_path(f"{save_path}\{file_time}")
    is_path(f"{save_path}\{file_time}\ip_static")

    find_ip = re.compile("(\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d?) ")
    ips = find_ip.findall(text)


    requests = text.split('\n')

    ip_dict = Counter(ips)

    ip_list = ip_dict.most_common(100)  # 统计访问次数前100的ip

    outjson = {"status:": "success", "result": []}
    outjson_ua = {"result": []}

    count_ip = {}
    with open(f'{save_path}\{file_time}\ip_static\ip.txt', 'a+') as f:
        for ip in ip_list:
            count_ip[ip[0]] = ip[1]
            outjson["result"].append({"count": ip[1], ip[0]: []})
            f.write(ip[0]+" "+str(ip[1])+"\n")

    # 请求正则
    qingqiu = re.compile('([A-Z]+ .*)" \d+ \d+')

    # 响应状态码
    head_code = re.compile('[A-Z]+ .* (\d+) \d+')

    # ua正则
    ua = re.compile('[A-Z]+ .* \d+ \d+.*?(\w+/.*)')

    # error
    error_qingqiu = re.compile('\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d? .*?"(.*?)" \d+ \d+')

    # error_code
    error_code = re.compile('\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d? .*?(\d+) \d+')

    # error_ua
    error_ua = re.compile('\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d?\.\d?\d?\d? .*?".*?" \d+ \d+.*?.*?(.*)')

    endlist = ""
    endua = ""
    for request in requests:
        find_qingqiu = qingqiu.findall(request)
        find_ua = ua.findall(request)
        find_head_code = head_code.findall(request)
        find_ip_2 = find_ip.findall(request)
        try:
            di = {find_qingqiu[0]: find_head_code[0]}

            text = find_ip_2[0] + find_qingqiu[0] + find_ua[0]
            for i in outjson["result"]:
                if find_ip_2[0] in i:
                    if text not in endlist:
                        endlist += text
                        i[find_ip_2[0]].append(di)
                        outjson_ua["result"].append([find_ip_2[0], find_qingqiu[0], find_ua[0]])

        except:
            find_error = error_qingqiu.findall(request)
            find_error_ua = error_ua.findall(request)
            find_error_code = error_code.findall(request)

            if request:
                error_di = {find_error[0]: find_error_code[0]}

                text_error = find_ip_2[0] + find_error[0] + find_error_ua[0]

                for i in outjson["result"]:
                    if find_ip_2[0] in i:
                        if text_error not in endua:
                            endua += text_error
                            i[find_ip_2[0]].append(error_di)
                            outjson_ua["result"].append([find_ip_2[0], find_error[0], find_error_ua[0].replace('\"', '')])

    if len(outjson["result"]) <=0:
        outjson["status:"]='error'

    save_excel(save_path,file_time,outjson,outjson_ua,ip_list)

    return outjson,outjson_ua,ip_list




#查询ip位置
def request(ip):
    apikey = "b54b66c16a25d938"

    url = "https://api.binstd.com/ip/location?appkey={}&ip={}"
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36',
        'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6',
    }

    newurl=url.format(apikey,ip)
    res = requests.get(url=newurl,headers=header)
    json = res.json()['result']
    country=json['country'] if json['country']!=None else "空"
    province =json['province'] if json['country']!=None else "空"
    city=json['city'] if json['country']!=None else "空"
    type =json['type'] if json['type']!=None else "空"
    return country,province,city,type

#列出ip
def iplist(save_path,file_time):
    with open(f"{save_path}\{file_time}\ip_static\ip.txt",'r') as f:
        ip_numbers=f.readlines()

    ip_dict={}
    with open(f"{save_path}\{file_time}\ip_static\\found_ip.txt",'a+') as f:

        for i in ip_numbers:
            ips=i.replace('\n','').split(' ')
            for i in range(len(ips)):
                ip_dict[ips[0]]=ips[1]
                f.write(ips[0]+"\n")

    return ip_dict

#查询ip地址
def where_ip(save_path,file_time):
    list=[]
    ip_dict = iplist(save_path,file_time)
    i=1
    j=1


    table = PrettyTable()
    table.field_names=["ip","国家","省份","城市","归属","访问次数"]

    for key in ip_dict:
        try:
            country,province,city,type=request(key)
            # print(key,country,province,city,ip_dict[key])
            list.append([key,country,province,city,type,ip_dict[key]])
            savetable(table,key,country,province,city,type,ip_dict[key])
            time.sleep(0.1)
        except:
            print(f"{key} is error")

    wb_ip=Workbook()
    ws_ip=wb_ip.active
    ws_ip.append(["ip","国家","省份","城市","归属","访问次数"])
    for i in list:
        ws_ip.append([i[0],i[1],i[2],i[3],i[4],i[5]])

    wb_ip.save(f"{save_path}\{file_time}\ip请求次数统计.xls")

